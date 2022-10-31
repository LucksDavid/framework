# coding = "UTF-8"
# 读取Excel文件
import openpyxl
import pytest


class Test_case1():
    def read_excel(excel_path, sheet_name, flag):
        # 加载工作簿
        wb = openpyxl.load_workbook(excel_path)
        # 获取sheet的对象
        sheet = wb[sheet_name]
        # 读取excel的数据，组成列表格式的数据
        row_list = []
        for row in range(2, sheet.max_row + 1):
            col_list = []
            if sheet.cell(row, flag).value == "y":
                for col in range(1, sheet.max_column + 1):
                    col_list.append(sheet.cell(row, col).value)
                row_list.append(col_list)
        return row_list

        @pytest.mark.parametrize("caseinfo", read_excel("../data/test_login.xlsx", "test_01_login", 5))
        def test_case1(caseinfo):
            print(caseinfo)


if __name__ == "__main__":
    pytest.main("test_api.py")
